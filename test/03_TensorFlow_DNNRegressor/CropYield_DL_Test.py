# -*- coding: utf-8 -*-
"""
Created on Tue Feb 23 16:13:21 2021

@author: Chutj
"""

# 加载必要的库、包等
import os
os.environ['TF_CPP_MIN_LOG_LEVEL']='3'
import openpyxl
import numpy as np
import pandas as pd
import tensorflow as tf
import scipy.stats as stats
import matplotlib.pyplot as plt
from sklearn import metrics
from sklearn.model_selection import train_test_split

# ===============*** 函数声明区域 ***===============

# DeleteOldModel函数，删除上一次运行所保存的模型
def DeleteOldModel(ModelPath):
    AllFileName=os.listdir(ModelPath) # 获取ModelPath路径下全部文件与文件夹
    for i in AllFileName:
        NewPath=os.path.join(ModelPath,i) # 分别将所获取的文件或文件夹名称与ModelPath路径组合
        if os.path.isdir(NewPath): # 若组合后的新路径是一个文件夹
            DeleteOldModel(NewPath) # 递归调用DeleteOldModel函数
        else:
            os.remove(NewPath) # 若不是一个新的文件夹，而是一个文件，那么就删除

# LoadData函数，加载全部数据
def LoadData(DataPath):
    MyData=pd.read_csv(DataPath,names=['EVI0610','EVI0626','EVI0712','EVI0728','EVI0813','EVI0829',
                                       'EVI0914','EVI0930','EVI1016','Lrad06','Lrad07','Lrad08',
                                       'Lrad09','Lrad10','Prec06','Prec07','Prec08','Prec09',
                                       'Prec10','Pres06','Pres07','Pres08','Pres09','Pres10',
                                       'SIF161','SIF177','SIF193','SIF209','SIF225','SIF241',
                                       'SIF257','SIF273','SIF289','Shum06','Shum07','Shum08',
                                       'Shum09','Shum10','SoilType','Srad06','Srad07','Srad08',
                                       'Srad09','Srad10','Temp06','Temp07','Temp08','Temp09',
                                       'Temp10','Wind06','Wind07','Wind08','Wind09','Wind10',
                                       'Yield'],header=0) # 加载DataPath路径所指定的数据，names中的内容为各列的名称
    return MyData

# InputFun函数，训练数据与验证数据所用的Input函数
def InputFun(Features,Labels,Training,BatchSize):
    Datasets=tf.data.Dataset.from_tensor_slices((dict(Features),Labels)) # 对数据加以加载
    if Training:
        Datasets=Datasets.shuffle(1000).repeat() # 对于训练数据，需要打乱（shuffle）、重复（repeat）
    return Datasets.batch(BatchSize) # 将经过上述处理后的数据以每次BatchSize个输出

# InputFunPredict函数，测试数据所用的Input函数
def InputFunPredict(Features,BatchSize):
    return tf.data.Dataset.from_tensor_slices(dict(Features)).batch(BatchSize) # 对数据加以加载,以每次BatchSize个输出

# AccuracyVerification函数，进行精度验证指标的计算与绘图
def AccuracyVerification(PredictLabels,TestLabels):
    value=0
    PredictValuesList=[]
    for k in PredictLabels:
        value=k.get('predictions')[0]
        PredictValuesList.append(value)
    TestLabels=TestLabels.values.tolist()
    TestYList=sum(TestLabels,[])
    # 以上为获取测试数据的因变量与模型预测所得的因变量
    Pearsonr=stats.pearsonr(TestYList,PredictValuesList) # 计算皮尔逊相关系数
    R2=metrics.r2_score(TestYList,PredictValuesList) # 计算R方
    RMSE=metrics.mean_squared_error(TestYList,PredictValuesList)**0.5 # 计算RMSE
    plt.cla()
    plt.plot(TestYList,PredictValuesList,'r*')
    plt.xlabel('Actual Values')
    plt.ylabel('Predicted Values')
    # 以上为绘制拟合图像
    print('Pearson correlation coefficient is {0}, and RMSE is {1}.'.format(Pearsonr[0],RMSE))
    return (Pearsonr[0],R2,RMSE,PredictValuesList)

# WriteAccuracy函数，将模型所涉及的参数与最终精度结果保存
def WriteAccuracy(*WriteVar):
    ExcelData=openpyxl.load_workbook(WriteVar[0])
    SheetName=ExcelData.get_sheet_names() # 获取全部Sheet
    WriteSheet=ExcelData.get_sheet_by_name(SheetName[0]) # 获取指定Sheet
    WriteSheet=ExcelData.active # 激活指定Sheet
    MaxRowNum=WriteSheet.max_row # 获取指定Sheet对应第一个空行
    for i in range(len(WriteVar)-1):
        exec("WriteSheet.cell(MaxRowNum+1,i+1).value=WriteVar[i+1]") # 用exec执行语句，写入信息
    ExcelData.save(WriteVar[0]) # 保存文件


# ===============*** 代码由此开始执行 ***===============
#      ++++++++++--- 建议由这里开始看 ---++++++++++

# 将各类变量放在一个位置集中定义，十分有利于机器学习等变量较多的代码
MyModelPath="G:/CropYield/03_DL/02_DNNModle" # 确定每一次训练所得模型保存的位置
MyDataPath="G:/CropYield/03_DL/00_Data/AllDataAll.csv" # 确定输入数据的位置
MyResultSavePath="G:/CropYield/03_DL/03_OtherResult/EvalResult54.xlsx" # 确定模型精度结果（RMSE等）与模型参数保存的位置
TestSize=0.2 # 确定数据中测试集所占比例
RandomSeed=np.random.randint(low=24,high=25) # 确定划分训练集与测试集的随机数种子
OptMethod='Adam' # 确定模型所用的优化方法
LearningRate=0.01 # 确定学习率
DecayStep=200 # 确定学习率下降的步数
DecayRate=0.96 # 确定学习率下降比率
HiddenLayer=[64,128] # 确定隐藏层数量与每一层对应的神经元数量
ActFun='tf.nn.relu' # 确定激活函数
Dropout=0.3 # 确定Dropout的值
LossReduction='tf.compat.v1.ReductionV2.SUM_OVER_BATCH_SIZE' # 指定每个批次训练误差的减小方法
BatchNorm='False' # 确定是否使用Batch Normalizing
TrainBatchSize=110 # 确定训练数据一个Batch的大小
TrainStep=3000 # 确定训练数据的Step数量
EvalBatchSize=1 # 确定验证数据一个Batch的大小
PredictBatchSize=1 # 确定预测数据（即测试集）一个Batch的大小

# 调用DeleteOldModel函数，删除上一次运行所保存的模型
DeleteOldModel(MyModelPath)

# 初始数据处理
AllXY=LoadData(MyDataPath) # 调用LoadData函数，获取数据
Label={"Yield":AllXY.pop("Yield")} # 将因变量从全部数据中提取出
AllX,AllY=AllXY,(pd.DataFrame(Label)) # 将自变量与因变量分离

# 划分数据训练集与测试集
TrainX,TestX,TrainY,TestY=train_test_split(AllX,
                                           AllY,
                                           test_size=TestSize, # 指定数据中测试集所占比例
                                           random_state=RandomSeed # 指定划分训练集与测试集的随机数种子
                                           )

# estimator接口中的模型需要用“Feature columns”对象作为输入数据，只有这样模型才知道读取哪些数据
FeatureColumn=[] # 定义一个新的“Feature columns”对象
for key in AllX.keys():
    FeatureColumn.append(tf.feature_column.numeric_column(key=key)) # 将全部因变量数据（需要均为连续变量）导入

# 定义模型优化方法
# Optimizer=OptMethod # 优化方法选用OptMethod所指定的方法
Optimizer=lambda:tf.keras.optimizers.Adam(
    learning_rate=tf.compat.v1.train.exponential_decay(learning_rate=LearningRate, # 初始学习率
                                                       global_step=tf.compat.v1.train.get_global_step(),
                                                       # 全局步数，用以计算已经衰减后的学习率
                                                       # get_global_step()函数自动获取当前的已经执行的步数
                                                       decay_steps=DecayStep, # 学习率下降完成的指定步数
                                                       decay_rate=DecayRate # 衰减率
                                                       ) # 选用基于学习率指数下降的Adam方法，此举有助于降低过拟合风险
                                                         # 这一函数返回每次对应的学习率
    )


# 基于DNNRegressor构建深度学习模型
DNNModel=tf.estimator.DNNRegressor(feature_columns=FeatureColumn, # 指定模型所用的“Feature columns”对象
                                   hidden_units=HiddenLayer, # 指定隐藏层数量与每一层对应的神经元数量
                                   optimizer=Optimizer, # 指定模型所用的优化方法                                  
                                   activation_fn=eval(ActFun), # 指定激活函数
                                   dropout=Dropout, # 指定Dropout的值
                                   label_dimension=1, # 输出数据的维度，即因变量的个数
                                   model_dir=MyModelPath, # 指定每一次训练所得模型保存的位置
                                   # loss_reduction=eval(LossReduction), # 指定每个批次训练误差的减小方法
                                   batch_norm=eval(BatchNorm) # 指定是否使用Batch Normalizing
                                   )

# tf.compat.v1.logging.set_verbosity(tf.compat.v1.logging.INFO) # 将INFO级别的日志信息显示到屏幕

# 基于训练数据训练模型
DNNModel.train(input_fn=lambda:InputFun(TrainX,
                                        TrainY,
                                        True,
                                        TrainBatchSize
                                        ), # 调用InputFun函数；InputFun函数返回“tf.data.Dataset”对象，这个对象才可以被
                                           # train函数识别并带入模型；由于InputFun函数每次返回BatchSize大小的数据个数，
                                           # 因此需要多次执行，前面需要加lambda
               steps=TrainStep # 指定模型训练的步数
               ) 

# 验证模型并保存验证结果
EvalResult=DNNModel.evaluate(input_fn=lambda:InputFun(TestX,
                                                      TestY,
                                                      False,
                                                      EvalBatchSize
                                                      )
                             )
# 打印验证结果
print('ev:{}'.format(EvalResult))

# 基于测试数据测试模型精度结果
PredictValues=DNNModel.predict(input_fn=lambda:InputFunPredict(TestX,
                                                               PredictBatchSize
                                                               )
                               )

# 调用AccuracyVerification函数，进行精度验证指标的计算与绘图
AccuracyResult=AccuracyVerification(PredictValues,TestY)
PearsonR,R2,RMSE,PredictY=AccuracyResult[0],AccuracyResult[1],AccuracyResult[2],AccuracyResult[3]

# 调用WriteAccuracy函数，将模型所涉及的参数与最终精度结果保存
WriteAccuracy(MyResultSavePath,PearsonR,R2,RMSE,TestSize,RandomSeed,OptMethod,LearningRate,DecayStep,
              DecayRate,','.join('%s' %i for i in HiddenLayer),ActFun,Dropout,LossReduction,
              BatchNorm,TrainBatchSize,TrainStep,EvalBatchSize,PredictBatchSize)